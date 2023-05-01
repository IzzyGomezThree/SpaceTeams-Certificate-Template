from openpyxl import load_workbook
from openpyxl.drawing.image import Image

from PIL import Image
from PIL import ImageDraw
from PIL import ImageFont

# get excel file and sheet
wb = load_workbook(filename = 'RealFiles/MasterFile.xlsx')
student_sheet = wb.get_sheet_names()[0]
worksheet = wb.get_sheet_by_name(student_sheet)
n_row = worksheet.max_row

for i in range (1, n_row + 1):
    # Open the raw Image
    img = Image.open('CertificateTemplates/studentCertificate_v2.png')

    # get size of photo, and declare the photo size and fontsize
    W, H = (1121, 793)

    # Call draw Method to add 2D graphics in an image
    draw = ImageDraw.Draw(img)

    # Read message from (as tuple)...
    firstName = worksheet.cell(row = i+1, column = 2)   # column 2, FirstName
    lastName = worksheet.cell(row = i+1, column = 3)   # column 3, LastName
    schoolName = worksheet.cell(row = i+1, column = 5)   # column 5, School

    # turn tuple variable to strings
    firstName_value = firstName.value
    lastName_value = lastName.value
    fullName_value = firstName_value + " " + lastName_value    # combine first and last name 
    schoolName_value = schoolName.value

    # Custom font style and font size
    arialFont = ImageFont.truetype("Fonts/arial.ttf", 30)

    # declare dynamic font size variable
    img_fraction = 0.80
    fontsize = 100
    
    # when the text goes out of bounds from the photo size, it slowly decreases until the text fills 95% of the photo)
    freehandFont = ImageFont.truetype("Fonts/Freehand-Regular.ttf", fontsize)
    while freehandFont.getsize(fullName_value)[0] > img_fraction*img.size[0]:
        # de-increment to be sure it is less than criteria
        fontsize -= 1
        freehandFont = ImageFont.truetype("Fonts/Freehand-Regular.ttf", fontsize)

    # declare line, (0,0) is top left corner
    y = 415     # the higher the value, the lower the line
    delta_x = 400   # the higher the value, the wider the line
    x1 = 560.5 - delta_x   # x-coordinate on the left side
    x2 = 560.5 + delta_x   # x-coordinate on the right side
    width_x  = x2-x1 
    shape = [(x1, y), (x2, y)]    # 560.5 is the midway point on the x-axix
    
    # dynamically increase the line size when the text surpasses the border by 60 pixels
    while freehandFont.getsize(fullName_value)[0] > (width_x-60):
        if width_x < 1064:      # 1064 pixels is the longest width the line is allowed to be
            delta_x += 1
            x1 = 560.5 - delta_x   # x-coordinate on the left side
            x2 = 560.5 + delta_x   # x-coordinate on the right side
            width_x  = x2-x1 
            shape = [(x1, y), (x2, y)]    # 560.5 is the midway point on the x-axix
        else:
            break

    # Add Text and line to an image
    draw.line(shape, fill=(0,0,0), width=6)
    draw.text((W/2,H/2-20), fullName_value, font=freehandFont, fill=(0, 0, 0), anchor="mm")
    draw.text((W/2,H/2+70), schoolName_value, font=arialFont, fill=(0, 0, 0), anchor="mm")

    # Save images
    img.save('C:/Users/Skullkrasher/Documents/GitHub/SpaceCraft-Certificates-Fall2022/Certificates/participationCertificates/'+firstName_value+'.'+lastName_value+'_Certificate.png', "PNG", resoultion=100.0)

    # add image file paths to column "G"
    filePath = worksheet.cell(row = i+1, column = 7)   # "G" is column #7
    filePath.value = 'C:/Users/Skullkrasher/Documents/GitHub/SpaceCraft-Certificates-Fall2022/Certificates/participationCertificates/'+firstName_value+'.'+lastName_value+'_Certificate.png'
    wb.save("C://Users//Skullkrasher//Documents//GitHub//SpaceCraft-Certificates-Fall2022//RealFiles//MasterFile.xlsx")

    # debug to see which certificate is done
    print(firstName_value, lastName_value )

print ("CERTIFICATES ARE DONE!!")