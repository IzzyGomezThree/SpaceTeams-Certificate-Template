# THIS IS A TEMPLATE, AND AS SUCH, MOST FILE PATHS ARE ABSENT AND NEED TO BE REPLACED
# FOLLOW ALONG WITH THE "Step-By-Step: Emailing Certificates" INSTRUCTIONS FOR UNDERSTANDING HOW TO USE THIS TEMPLATE

# If you are planning to email vouchers:
# Uncomment all lines data processing lines (37, 43, 53). 
# Change the email HTML template to include a comment about the voucher.

# ALL FILE PATHS WORK IF FOLDER/FILE ORGANIZATION IS SETUP PER THE STEP-BY-STEP INSTRUCTIONS
# IF A FILE CANNOT BE FOUND, OR YOU WANT TO CHANGE THE FILE PATH, MAKE SURE TO COPY THE ENTIRE FILE PATH OF THE FILE
# EXAMPLE: RealFiles/MasterFile.xlsx = C:/Users/Skullkrasher/Documents/GitHub/SpaceCraft-Certificates-Fall2022/RealFiles/MasterFile.xlsx

# STEP 2
# LIBRARY INSTALL GUIDE
# To install redmail, type "pip install redmail" in the python command line
# To install OpenPyxl, type "pip install openpyxl" in the python command line (if not already installed)
# To install pathlib, type "pip install pathlib" in the python command line (Note that if you're using Python 3.4 or above, pathlib pathlib is now part of the standard library, and should already be installed)

# STEP 3
# Declare libraries
from redmail import gmail
from pathlib import Path
from openpyxl import load_workbook

# STEP 4
# get excel file and sheet
# wb = load_workbook(filename = 'RealFiles/MasterFile.xlsx') # THIS IS COMMENTED OUT TO PREVENT ACCIDENTAL EXECUTION!!!!!! ONLY UNCOMMENT WHEN READY TO EXECUTE!!!!! 
student_sheet = wb.get_sheet_names()[0]
worksheet = wb.get_sheet_by_name(student_sheet)
n_row = worksheet.max_row

# STEP 5
for i in range (1, n_row + 1):
    
    # STEP 6
    # Read data from excel (as tuple)...
    emailName = worksheet.cell(row = i+1, column = 6)   # column 6 = column F, Email
    fileLocationAward = worksheet.cell(row = i+1, column = 4)   # column 4 = column D, File Location   
    # fileLocationVoucher = worksheet.cell(row = i+1, column = 5)   # column 5 = column E, File Location   

    # STEP 7
    # turn tuple variable to strings
    emailName_value = emailName.value
    fileLocationAward_value = fileLocationAward.value
    # fileLocationVoucher_value = fileLocationVoucher.value
    
    # STEP 8
    # if there is an empty cell for the award, continue to the next row
    if fileLocationAward_value is None:
        continue

    # STEP 9
    # replace the long file name to just the file name
    fileName_award = fileLocationAward_value.replace('C:/Users/Skullkrasher/Documents/GitHub/SpaceCraft-Certificates-Fall2022/Certificates/awardCertificates/', '')
    # fileName_voucher = fileLocationVoucher_value.replace('C:/Users/Skullkrasher/Documents/GitHub/SpaceCraft-Certificates-Fall2022/Certificates/vouchersPNG/', '')

    # STEP 10 - Spacecraft
    # log into Gmail account
    gmail.username = 'spacecraft.vr@gmail.com'
    gmail.password = 'passwordHere'

    # STEP 10 - Debugging
    # # gmail account debugger
    # gmail.username = 'urEmail@here.com'
    # gmail.password = 'passwordHere'

    # STEP 11
    # And then you can send emails
    gmail.send(
        
        # STEP 12
        subject="Space Teams - Mission Oz 2022 Award Certificate",
        sender="Space Teams <spacecraft.vr@gmail.com>",
        # bcc=['izzygomezthree@gmail.com', 'izzygomezthree@tamu.edu'],
        receivers=[emailName_value],
        attachments={
            fileName_award: Path(fileLocationAward_value),
            fileName_voucher: Path(fileLocationVoucher_value)
            },
        
        # STEP 13
        html=
        """
        <div class="gmail_default" style="font-size:large">Dear Space Teams cadets,</div>
        <div class="gmail_default" style="font-size:large"><br></div>
        <div class="gmail_default" style="font-size:large">Congratulations once again, and thank you for your participating in Mission Oz 2022. </div>
        <div class="gmail_default" style="font-size:large"><br></div>
        <div class="gmail_default" style="font-size:large">Attached to this email is your award certificate. If you feel any changes need to be made to the certificate, do not hesitate to reply to this email.</div>
        <div class="gmail_default" style="font-size:large"><br></div>
        <div class="gmail_default" style="font-size:large">Ad Astra</div>
        <div class="gmail_default" style="font-size:large"><br></div>
        {{SpaceTeamsLogo}}
        """,
        body_images={'SpaceTeamsLogo': 'EmailMedia/spaceTeams_email.png'}
    )

    # STEP 14
    print("Email Sent!")
