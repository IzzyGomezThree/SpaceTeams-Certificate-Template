from redmail import gmail
from pathlib import Path

from openpyxl import load_workbook

# get excel file and sheet
# wb = load_workbook(filename = 'RealFiles/MasterFile.xlsx') # THIS IS COMMENTED OUT TO PREVENT ACCIDENTAL EXECUTION!!!!!! ONLY UNCOMMENT WHEN READY TO EXECUTE!!!!! 
student_sheet = wb.get_sheet_names()[0]
worksheet = wb.get_sheet_by_name(student_sheet)
n_row = worksheet.max_row

for i in range (1, n_row + 1):
    # Read data from excel (as tuple)...
    emailName = worksheet.cell(row = i+1, column = 4)   # column 4 = column D, Email
    fileLocationAward = worksheet.cell(row = i+1, column = 8)   # column 8 = column H, File Location   
    fileLocationVoucher = worksheet.cell(row = i+1, column = 9)   # column 9 = column I, File Location   

    # turn tuple variable to strings
    emailName_value = emailName.value
    fileLocationAward_value = fileLocationAward.value
    fileLocationVoucher_value = fileLocationVoucher.value
    
    # if there is an empty cell for the award, continue to the next row
    if fileLocationAward_value is None:
        continue

    # replace the long file name to just the file name
    fileName_award = fileLocationAward_value.replace('C:/Users/Skullkrasher/Documents/GitHub/SpaceCraft-Certificates-Fall2022/Certificates/awardCertificates/', '')
    fileName_voucher = fileLocationVoucher_value.replace('C:/Users/Skullkrasher/Documents/GitHub/SpaceCraft-Certificates-Fall2022/Certificates/vouchersPNG/', '')

    # log into Gmail account
    gmail.username = 'spacecraft.vr@gmail.com'
    gmail.password = ''

    # # gmail account debugger
    # gmail.username = ''
    # gmail.password = ''

    # And then you can send emails
    gmail.send(
        subject="Space Teams - Mission Oz 2022 Award Certificate",
        sender="Space Teams <spacecraft.vr@gmail.com>",
        # bcc=['izzygomezthree@gmail.com', 'izzygomezthree@tamu.edu'],
        receivers=[emailName_value],
        attachments={
            fileName_award: Path(fileLocationAward_value),
            fileName_voucher: Path(fileLocationVoucher_value)
            },
        html=
        """
        <div class="gmail_default" style="font-size:large">Dear Space Teams Winners,</div>
        <div class="gmail_default" style="font-size:large"><br></div>
        <div class="gmail_default" style="font-size:large">Congratulations once again, and thank you for your participating in Mission Oz 2022. </div>
        <div class="gmail_default" style="font-size:large"><br></div>
        <div class="gmail_default" style="font-size:large">Attached to this email is your award certificate and your voucher. If you feel any changes need to be made to the certificate or the voucher, do not hesitate to reply to this email. Also reply if the voucher does not work.</div>
        <div class="gmail_default" style="font-size:large"><br></div>
        <div class="gmail_default" style="font-size:large">Ad Astra</div>
        <div class="gmail_default" style="font-size:large"><br></div>
        {{SpaceTeamsLogo}}
        """,
        body_images={'SpaceTeamsLogo': 'EmailMedia/spaceTeams_email.png'}
    )

    print("Email Sent!")
