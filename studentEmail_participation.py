from redmail import gmail
from pathlib import Path

from openpyxl import load_workbook

# get excel file and sheet
# wb = load_workbook(filename = 'RealFiles/MasterFile.xlsx') # THIS IS COMMENTED OUT TO PREVENT ACCIDENTAL EXECUTION!!!!!! ONLY UNCOMMENT WHEN READY TO EXECUTE!!!!! 
student_sheet = wb.get_sheet_names()[0]
worksheet = wb.get_sheet_by_name(student_sheet)
n_row = worksheet.max_row

for i in range (1, n_row + 1):
    # Read data from excel (as tuple)...
    emailName = worksheet.cell(row = i+1, column = 4)   # column 4 = column D, Email
    fileLocation = worksheet.cell(row = i+1, column = 7)   # column 7 = column G, File Location

    # turn tuple variable to strings
    emailName_value = emailName.value
    fileLocation_value = fileLocation.value
    fileName = fileLocation_value.replace('C:/Users/Skullkrasher/Documents/GitHub/SpaceCraft-Certificates-Fall2022/Certificates/participationCertificates/', '')

    # log into Gmail account
    gmail.username = 'spacecraft.vr@gmail.com'
    gmail.password = ''

    # # gmail account debugger
    # gmail.username = 'izzygomezthree@gmail.com'
    # gmail.password = ''

    # And then you can send emails
    gmail.send(
        subject="Space Teams - Mission Oz 2022 Participation Certificate",
        sender="Space Teams <spacecraft.vr@gmail.com>",
        # bcc=['izzygomezthree@gmail.com', 'izzygomezthree@tamu.edu'],
        receivers=[emailName_value],
        attachments={fileName: Path(fileLocation_value)},
        html=
        """
        <div class="gmail_default" style="font-size:large">Dear Space Cadets,</div>
        <div class="gmail_default" style="font-size:large"><br></div>
        <div class="gmail_default" style="font-size:large">Congratulations again on your participation in Mission Oz 2022! Some of you achieved the highest scores ever, and we hope you also learned a great deal about conducting a successful space mission. Thank you for your enthusiasm and determination, and we hope you continue your exploration of the stars.</div>
        <div class="gmail_default" style="font-size:large"><br></div>
        <div class="gmail_default" style="font-size:large">Attached to this email is your Space Teams certificate. If you feel any changes need to be made, do not hesitate to reply to this email. If you were on a winning team, then prizes are coming.</div>
        <div class="gmail_default" style="font-size:large"><br></div>
        <div class="gmail_default" style="font-size:large">Ad Astra</div>
        <div class="gmail_default" style="font-size:large"><br></div>
        {{SpaceTeamsLogo}}
        """,
        body_images={'SpaceTeamsLogo': 'EmailMedia/spaceTeams_email.png'}
    )

    print("Email Sent!")
