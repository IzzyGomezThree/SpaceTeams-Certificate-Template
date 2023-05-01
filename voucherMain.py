from openpyxl import load_workbook
from openpyxl.drawing.image import Image

from PIL import Image
from PIL import ImageDraw
from PIL import ImageFont

# get excel file and sheet
wb = load_workbook(filename = 'RealFiles/MasterFile.xlsx')
student_sheet = wb.get_sheet_names()[0]
worksheet = wb.get_sheet_by_name(student_sheet)
n_row = worksheet.max_row

for j in range (1, 140):
    # Open the raw Image
    img = Image.open('CertificateTemplates/Voucher.png')

    # get size of photo, and declare the photo size and fontsize
    # print(img.size)
    W, H = (2000, 1410)

    # Call draw Method to add 2D graphics in an image
    draw = ImageDraw.Draw(img)

    # Read message from (as tuple)...
    scoreAward = worksheet.cell(row = j+1, column = 1)   # column 1 = column A, Score
    firstName = worksheet.cell(row = j+1, column = 2)   # column 2 = column B, FirstName
    lastName = worksheet.cell(row = j+1, column = 3)   # column 3 = column C, LastName
    voucherCode = worksheet.cell(row = j+1, column = 6)   # column #F, Voucher Code

    # turn tuple variable to strings
    scoreAward_value = scoreAward.value
    firstName_value = firstName.value
    lastName_value = lastName.value
    voucherCode_value = voucherCode.value

    # go through empty cells until there is a value; if there is a value, split up the words and look to 1st, 2nd, or 3rd, and then assign the respective values
    if scoreAward_value is None:    # "None" = cell empty, so continue to the next row
        continue
    else:
        split = scoreAward_value.split()    # split the sentence into individual words
        if '1st' in split:  # if the word "1st" is found, assign the following values
            voucher_value = '100'
            voucher_text = '$100'
        elif '2nd' in split:    # if the word "2nd" is found, assign the following values
            voucher_value = '75'
            voucher_text = '$75'
        elif '3rd' in split:    # if the word "3rd" is found, assign the following values
            voucher_value = '50'
            voucher_text = '$50'
    
    # Custom font style and font size
    opensansFont_small = ImageFont.truetype("Fonts/open-sans.semibold.ttf", 60)
    opensansFont_voucherAmount = ImageFont.truetype("Fonts/open-sans.semibold.ttf", 115)
    opensansFont_voucherCode = ImageFont.truetype("Fonts/open-sans.semibold.ttf", 100)

    # Add Text and line to an image
    draw.text((W/2,H/2-250), scoreAward_value + " Winner", font=opensansFont_small, fill=(255, 255, 255), features=None, anchor="mm")   # draw.text((width & height), wording, font, fill color, features, anchor (which is the middle[mm]))
    draw.text((W/2+0,H/2+40), voucher_text, font=opensansFont_voucherAmount, fill=(255, 255, 255), features=None, anchor="mm")
    draw.text((W/2+0,H/2+560), voucherCode_value, font=opensansFont_voucherCode, fill=(255, 255, 255), features=None, anchor="mm")

    # save the modified image
    img.save('C:/Users/Skullkrasher/Documents/GitHub/SpaceCraft-Certificates-Fall2022/Certificates/vouchersPNG/'+firstName_value+'.'+lastName_value+'_Voucher'+voucher_value+'.png', "PNG", resoultion=100.0)

    # add PNG file paths to column "I"
    filePath_PNG = worksheet.cell(row = j+1, column = 9)   # "I" is column #9
    filePath_PNG.value = 'C:/Users/Skullkrasher/Documents/GitHub/SpaceCraft-Certificates-Fall2022/Certificates/vouchersPNG/'+firstName_value+'.'+lastName_value+'_Voucher'+voucher_value+'.png'
    wb.save("C://Users//Skullkrasher//Documents//GitHub//SpaceCraft-Certificates-Fall2022//RealFiles//MasterFile.xlsx")

    # # convert PNG to PDF
    # PNG = Image.open('C:/Users/Skullkrasher/Documents/GitHub/Certificate-Generator/Certificates/vouchers/'+firstName_value+'.'+lastName_value+'_Voucher'+voucher_value+'.png')
    # PDF = PNG.convert('RGB')
    # PDF.save('C:/Users/Skullkrasher/Documents/GitHub/Certificate-Generator/Certificates/voucherPDF/'+firstName_value+'.'+lastName_value+'_Voucher'+voucher_value+'.pdf', "PDF", resoultion=100.0)

    # # save PDF file name to excel
    # filePath_PDF = worksheet.cell(row = j+1, column = 39)   # "AM" is column #39
    # filePath_PDF.value = 'C:/Users/Skullkrasher/Documents/GitHub/Certificate-Generator/Certificates/voucherPDF/'+firstName_value+'.'+lastName_value+'_Voucher'+voucher_value+'.pdf'
    # wb.save("C://Users//Skullkrasher//Documents//GitHub//Certificate-Generator//RealFiles//Master_Prize_List.xlsx")

    # debug to see which certificate is done
    print(firstName_value, lastName_value, "Voucher" )
    
print(" Vouchers done")
