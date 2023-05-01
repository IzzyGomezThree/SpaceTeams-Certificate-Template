# THIS IS A TEMPLATE, AND AS SUCH, MOST FILE PATHS ARE ABSENT AND NEED TO BE REPLACED
# FOLLOW ALONG WITH THE "Step-By-Step: Certificate Creation" INSTRUCTIONS FOR UNDERSTANDING HOW TO USE THIS TEMPLATE

# ALL FILE PATHS WORK IF FOLDER/FILE ORGANIZATION IS SETUP PER THE STEP-BY-STEP INSTRUCTIONS
# IF A FILE CANNOT BE FOUND, OR YOU WANT TO CHANGE THE FILE PATH, MAKE SURE TO COPY THE ENTIRE FILE PATH OF THE FILE
# EXAMPLE: RealFiles/MasterFile.xlsx = C:/Users/Skullkrasher/Documents/GitHub/SpaceCraft-Certificates-Fall2022/RealFiles/MasterFile.xlsx

# STEP 2
# LIBRARY INSTALL GUIDE
# To install Pillow, type "pip install Pillow" in the python command line
# To install OpenPyxl, type "pip install openpyxl" in the python command line

# STEP 3
# Declare libraries
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

from PIL import Image
from PIL import ImageDraw
from PIL import ImageFont

# STEP 4 
# Get excel workbook and worksheet
wb = load_workbook(filename = 'RealFiles/Template-Excel-File.xlxs')   
student_sheet = wb.get_sheet_names()[0]     # "[0]" means this is the first worksheet; "[1]" would mean the second worksheet
worksheet = wb.get_sheet_by_name(student_sheet)
n_row = worksheet.max_row

# STEP 5
# This is the code that loops for every row
for i in range (1, n_row + 1):
    # STEP 6
    # Open the raw Image
    img = Image.open('CertificateTemplates/studentCertificate_v2.png')

    # STEP 7
    # get size of photo, and declare the photo size
    # W, H = img.size  # you can get the image resolution with "img.size"
    W, H = (1121, 793)  

    # STEP 8
    # Call draw Method to add 2D graphics in an image
    draw = ImageDraw.Draw(img)

    # STEP 9
    # Read values from excel sheet (as tuple)...
    firstName = worksheet.cell(row = i+1, column = 1)   # column 1, FirstName
    lastName = worksheet.cell(row = i+1, column = 2)   # column 2, LastName
    schoolName = worksheet.cell(row = i+1, column = 3)   # column 3, School

    # STEP 10
    # turn tuple variable to strings
    firstName_value = firstName.value
    lastName_value = lastName.value
    fullName_value = firstName_value + " " + lastName_value    # combine first and last name 
    schoolName_value = schoolName.value
    
    # STEP 11
    # Declare custom font style and font size
    arialFont = ImageFont.truetype("Fonts/arial.ttf", 30)  

    # STEP 12
    # Declare dynamic font size variable
    # fontsize is the starting font size that will get smaller in Step 14 
    img_fraction = 0.80
    fontsize = 100
    
    # STEP 13
    # when the text goes out of bounds from the photo size, it slowly decreases until the text fills 80% of the photo)
    # the percentage can be changed with the "img_fraction" variable
    freehandFont = ImageFont.truetype("Fonts/Freehand-Regular.ttf", fontsize)
    while freehandFont.getsize(fullName_value)[0] > img_fraction*img.size[0]:
        # de-increment to be sure it is less than criteria
        fontsize -= 1
        freehandFont = ImageFont.truetype("Fonts/Freehand-Regular.ttf", fontsize)

    # STEP 14
    # declare line; (0,0) is top left corner
    y = 415     # the higher the value, the lower the line
    delta_x = 400   # the higher the value, the wider the line
    x1 = 560.5 - delta_x   # x-coordinate on the left side
    x2 = 560.5 + delta_x   # x-coordinate on the right side
    width_x  = x2-x1 
    shape = [(x1, y), (x2, y)]    # 560.5 is the midway point on the x-axix
    
    # STEP 15
    # dynamically increase the line size when the text surpasses the border by 60 pixels
    while freehandFont.getsize(fullName_value)[0] > (width_x-60):
        if width_x < 1064:      # 1064 pixels is the longest width the line is allowed to be
            delta_x += 1
            x1 = 560.5 - delta_x   # x-coordinate on the left side
            x2 = 560.5 + delta_x   # x-coordinate on the right side
            width_x = x2-x1 
            shape = [(x1, y), (x2, y)]    # 560.5 is the midway point on the x-axix
        else:
            break

    # STEP 16
    # Add Text and line to an image
    draw.line(shape, fill=(0,0,0), width=6)
    draw.text((W/2,H/2-20), fullName_value, font=freehandFont, fill=(0, 0, 0), anchor="mm")
    draw.text((W/2,H/2+70), schoolName_value, font=arialFont, fill=(0, 0, 0), anchor="mm")

    # STEP 17
    # Save images as "firstName.lastName_AwardCertificate" under the "awardCertificates" folder; for example, the certificate should save as "Israel.Gomez_awardCertificate"
    # img.save('C:/Users/Skullkrasher/Documents/GitHub/SpaceCraft-Certificates-Fall2022/Certificates/awardCertificates/'+firstName_value+'.'+lastName_value+'_AwardCertificate.png', "PNG", resoultion=100.0)
    # example above^
    img.save('insertFilePathHere/'+firstName_value+'.'+lastName_value+'_AwardCertificate.png', "PNG", resoultion=100.0)

    # STEP 18
    # add image file paths to column "D"
    filePath = worksheet.cell(row = i+1, column = D)   # "D" is column #4
    filePath.value = 'C:/Users/Skullkrasher/Documents/GitHub/SpaceCraft-Certificates-Fall2022/Certificates/awardCertificates/'+firstName_value+'.'+lastName_value+'_AwardCertificate.png'
    wb.save("C://Users//Skullkrasher//Documents//GitHub//SpaceCraft-Certificates-Fall2022//RealFiles//MasterFile.xlsx")

    # STEP 19
    # debug to see which certificate is done
    print(firstName_value, lastName_value, "Award Certificate" )
print(" Score Loop done")
